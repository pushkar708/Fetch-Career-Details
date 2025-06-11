from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import random
from openpyxl import Workbook
import time
from openpyxl import load_workbook
import pyautogui as py
import os

cwd = os.path.dirname(os.path.abspath(__file__))


def return_final_url(url,page_number):
    import re
    final=url
    url_parts=final.split("?")
    if not re.search(r'\d$', url_parts[0]):
        final = f"{url_parts[0]}-{page_number}?{url_parts[1]}"
    return final


try:
    wb = load_workbook(os.path.join(cwd,"job_information.xlsx"))
    ws = wb.active
except FileNotFoundError:
    # If the file doesn't exist, create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.append(["Job Name", "Provider", "Experience Required", "Salary", "Location", "Description", "Qualifications", "Posted On", "Job URL"])

chrome_options=Options()
# chrome_options.add_argument("--headless")
driver=webdriver.Chrome(options=chrome_options)
for page_number in range(1, 7):
    temp_url = f"https://www.naukri.com/software-developer-jobs?k=software%20developer&nignbevent_src=jobsearchDeskGNB&experience=2&ctcFilter=6to10&jobAge=15"
    url=return_final_url(temp_url,page_number)
    driver.get(url)
    time.sleep(7)

    job_names_driver = driver.find_elements(By.XPATH, "//div[@class='srp-jobtuple-wrapper']//div[contains(@class,'row1')]/a")
    job_providers_driver = driver.find_elements(By.XPATH, "//div[@class='srp-jobtuple-wrapper']//div[contains(@class,'row2')]//a[contains(@class,'comp-name')]")
    job_details_driver = driver.find_elements(By.XPATH, "//div[@class='srp-jobtuple-wrapper']//div[contains(@class,'row3')]//span[contains(@class,'ni-job-tuple-icon')]")
    job_desc_driver=driver.find_elements(By.XPATH,"//div[@class='srp-jobtuple-wrapper']//div[contains(@class,'row4')]//span[contains(@class,'ni-job-tuple-icon')]")
    job_qualifications = driver.find_elements(By.XPATH,"//div[@class='srp-jobtuple-wrapper']//div[contains(@class,'row5')]//ul")
    job_posted_driver = driver.find_elements(By.XPATH,"//div[@class='srp-jobtuple-wrapper']//div[contains(@class,'row6')]//span[contains(@class,'job-post-day')]")

    # Loop through job_details_driver and add data to the Excel workbook
    for i in range(0, len(job_details_driver), 3):
        job_name = job_names_driver[i // 3].text if i // 3 < len(job_names_driver) else ""
        job_url = job_names_driver[i // 3].get_attribute("href") if i // 3 < len(job_names_driver) else ""
        provider = job_providers_driver[i // 3].text if i // 3 < len(job_providers_driver) else ""
        posted_on = job_posted_driver[i // 3].text if i // 3 < len(job_posted_driver) else ""
        desc = job_desc_driver[i // 3].text if i // 3 < len(job_desc_driver) else ""
        req_exp = job_details_driver[i].text if i < len(job_details_driver) else ""
        salary = job_details_driver[i + 1].text if i + 1 < len(job_details_driver) else ""
        location = job_details_driver[i + 2].text if i + 2 < len(job_details_driver) else ""

        # Handle job qualifications for each job posting
        qualifications = job_qualifications[i // 3] if i // 3 < len(job_qualifications) else None
        qualification_texts = []
        if qualifications:
            qualification_texts = [li.text for li in qualifications.find_elements(By.TAG_NAME, 'li')]
        joined_qualifications = ", ".join(qualification_texts)

        ws.append([job_name, provider, req_exp, salary, location, desc, joined_qualifications,posted_on,job_url])
# Save the workbook
wb.save(os.path.join(cwd,"job_information.xlsx"))






driver.quit()


